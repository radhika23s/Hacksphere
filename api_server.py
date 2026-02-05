#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
HealthifyAI Flask API Server
Serves ML predictions and admin dashboard
"""

from flask import Flask, request, jsonify, send_from_directory, session
from flask_cors import CORS
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from models import db, User
import os
import json
import pandas as pd
import numpy as np
from sklearn.tree import DecisionTreeClassifier
from sklearn.preprocessing import MultiLabelBinarizer
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score
import difflib
import warnings
from openpyxl import Workbook
from datetime import datetime

warnings.filterwarnings("ignore")

# ========== Setup Data Paths ==========
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

app = Flask(__name__)
app.config['SECRET_KEY'] = 'your-secret-key-here-change-in-production'
app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{os.path.join(BASE_DIR, "instance", "healthifyai.db")}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

CORS(app, resources={r"/api/*": {"origins": "*"}}, supports_credentials=True)

# Initialize extensions
db.init_app(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))
DATA_DIR = BASE_DIR

FILES = {
    'dataset': os.path.join(DATA_DIR, 'dataset.csv'),
    'severity': os.path.join(DATA_DIR, 'Symptom-severity.csv'),
    'description': os.path.join(DATA_DIR, 'symptom_Description.csv'),
    'precaution': os.path.join(DATA_DIR, 'symptom_precaution.csv'),
    'workout': os.path.join(DATA_DIR, 'gym dataset.csv')
}

for key, path in FILES.items():
    if not os.path.exists(path):
        print(f"⚠️  Missing {key}: {path}")

# ========== Utility Functions ==========
def normalize(sym):
    return sym.strip().lower().replace(" ", "_")

# Load datasets
try:
    dataset = pd.read_csv(FILES['dataset'])
    severity = pd.read_csv(FILES['severity'])
    description = pd.read_csv(FILES['description'])
    precaution = pd.read_csv(FILES['precaution'])
    workout = pd.read_csv(FILES['workout'])
    print("✅ All datasets loaded successfully")
except Exception as e:
    print(f"❌ Error loading datasets: {e}")
    exit(1)

# Train model
symptom_cols = [col for col in dataset.columns if "Symptom_" in col]
dataset["Symptoms"] = dataset[symptom_cols].apply(
    lambda row: [normalize(s) for s in row.dropna().values], axis=1
)

mlb = MultiLabelBinarizer()
X = mlb.fit_transform(dataset["Symptoms"])
y = dataset["Disease"]
known_symptoms = set(mlb.classes_)

X_train, X_test, y_train, y_test = train_test_split(
    X, y, test_size=0.2, random_state=42, stratify=y
)

clf = DecisionTreeClassifier(random_state=42)
clf.fit(X_train, y_train)
model_accuracy = accuracy_score(y_test, clf.predict(X_test))
print(f"✅ Model Accuracy: {model_accuracy * 100:.2f}%")

severity_dict = {
    normalize(sym): wt
    for sym, wt in zip(severity["Symptom"], severity["weight"])
}

def match_symptom(sym):
    matches = difflib.get_close_matches(sym, known_symptoms, n=1, cutoff=0.6)
    return matches[0] if matches else None

def predict_disease(user_symptoms):
    user_input = mlb.transform([user_symptoms])
    pred = clf.predict(user_input)[0]
    risk_score = sum(severity_dict.get(sym, 1) for sym in user_symptoms)
    
    try:
        desc = description.loc[description["Disease"] == pred, "Description"].values[0]
    except:
        desc = "No description available."
    
    try:
        prec = precaution.loc[
            precaution["Disease"] == pred,
            ["Precaution_1", "Precaution_2", "Precaution_3"]
        ].values[0]
        prec_str = ", ".join([p for p in prec if pd.notnull(p)])
    except:
        prec_str = "No specific precautions found."
    
    return pred, risk_score, desc, prec_str

def get_fitness_plan(age, gender, height_cm, weight_kg, goal, hypertension):
    sex = "Male" if str(gender).strip().lower() in ["male", "m"] else "Female"
    height_m = float(height_cm) / 100
    bmi = float(weight_kg) / (height_m ** 2)
    hyper_str = "Yes" if str(hypertension).strip().lower() == "yes" else "No"

    goal_map = {
        "lose weight": "Weight Loss",
        "gain muscle": "Weight Gain",
        "stay fit": "Weight Gain"
    }

    user_goal = str(goal).strip().lower()
    if user_goal == "stay fit":
        target_goal = "Weight Loss" if bmi >= 25 else "Weight Gain"
    else:
        target_goal = goal_map.get(user_goal, "Weight Gain")

    df_filtered = workout[
        (workout['Sex'] == sex) &
        (workout['Hypertension'] == hyper_str) &
        (workout['Fitness Goal'] == target_goal)
    ].copy()

    if df_filtered.empty:
        df_filtered = workout[
            (workout['Sex'] == sex) &
            (workout['Fitness Goal'] == target_goal)
        ].copy()

    if df_filtered.empty:
        df_filtered = workout[workout['Fitness Goal'] == target_goal].copy()

    if df_filtered.empty:
        return "No plan available", "No exercises available", "Consult a fitness professional"

    df_filtered['dist'] = (
        ((df_filtered['Age'] - age) / 100) ** 2 +
        ((df_filtered['BMI'] - bmi) / 50) ** 2
    )

    best_match = df_filtered.loc[df_filtered['dist'].idxmin()]
    return best_match.get('Diet', ''), best_match.get('Exercises', ''), best_match.get('Recommendation', '')

def save_user_to_excel(user_data):
    """Save user registration data to Excel file"""
    excel_file = os.path.join(BASE_DIR, 'user_registrations.xlsx')

    # Prepare data
    data = {
        'Registration Date': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'Username': user_data.get('username', ''),
        'Email': user_data.get('email', ''),
        'Age': user_data.get('age', ''),
        'Gender': user_data.get('gender', ''),
        'Height (cm)': user_data.get('height_cm', ''),
        'Weight (kg)': user_data.get('weight_kg', ''),
        'Goal': user_data.get('goal', ''),
        'Hypertension': user_data.get('hypertension', '')
    }

    try:
        # Check if file exists
        if os.path.exists(excel_file):
            # Load existing workbook
            from openpyxl import load_workbook
            wb = load_workbook(excel_file)
            ws = wb.active
        else:
            # Create new workbook
            wb = Workbook()
            ws = wb.active
            # Add headers if new file
            headers = list(data.keys())
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=1, column=col_num, value=header)

        # Find next empty row
        next_row = ws.max_row + 1

        # Add data
        for col_num, value in enumerate(data.values(), 1):
            ws.cell(row=next_row, column=col_num, value=value)

        # Save workbook
        wb.save(excel_file)
        print(f"✅ User data saved to {excel_file}")
        return True
    except Exception as e:
        print(f"❌ Error saving to Excel: {e}")
        return False

# ========== Authentication Endpoints ==========

@app.route('/api/auth/register', methods=['POST'])
def register():
    """Register a new user"""
    try:
        data = request.get_json(force=True)
        username = data.get('username', '').strip()
        email = data.get('email', '').strip()
        password = data.get('password', '')

        if not username or not email or not password:
            return jsonify({'error': 'Username, email, and password are required'}), 400

        if User.query.filter_by(username=username).first():
            return jsonify({'error': 'Username already exists'}), 400

        if User.query.filter_by(email=email).first():
            return jsonify({'error': 'Email already exists'}), 400

        user = User(username=username, email=email)
        user.set_password(password)

        db.session.add(user)
        db.session.commit()

        # Save user data to Excel
        save_user_to_excel(user.to_dict())

        login_user(user)
        return jsonify({'message': 'Registration successful', 'user': user.to_dict()})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

@app.route('/api/auth/login', methods=['POST'])
def login():
    """Login user"""
    try:
        data = request.get_json(force=True)
        username = data.get('username', '').strip()
        password = data.get('password', '')

        if not username or not password:
            return jsonify({'error': 'Username and password are required'}), 400

        user = User.query.filter_by(username=username).first()
        if not user or not user.check_password(password):
            return jsonify({'error': 'Invalid username or password'}), 401

        login_user(user)
        return jsonify({'message': 'Login successful', 'user': user.to_dict()})
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/auth/logout', methods=['POST'])
@login_required
def logout():
    """Logout user"""
    logout_user()
    return jsonify({'message': 'Logout successful'})

@app.route('/api/auth/status', methods=['GET'])
def auth_status():
    """Check authentication status"""
    if current_user.is_authenticated:
        return jsonify({'authenticated': True, 'user': current_user.to_dict()})
    return jsonify({'authenticated': False})

# ========== Profile Endpoints ==========

@app.route('/api/profile', methods=['GET'])
@login_required
def get_profile():
    """Get user profile"""
    return jsonify(current_user.to_dict())

@app.route('/api/profile', methods=['PUT'])
@login_required
def update_profile():
    """Update user profile"""
    try:
        data = request.get_json(force=True)

        # Update profile fields
        if 'age' in data:
            current_user.age = int(data['age']) if data['age'] else None
        if 'gender' in data:
            current_user.gender = data['gender']
        if 'height_cm' in data:
            current_user.height_cm = float(data['height_cm']) if data['height_cm'] else None
        if 'weight_kg' in data:
            current_user.weight_kg = float(data['weight_kg']) if data['weight_kg'] else None
        if 'goal' in data:
            current_user.goal = data['goal']
        if 'hypertension' in data:
            current_user.hypertension = data['hypertension']

        db.session.commit()
        return jsonify({'message': 'Profile updated successfully', 'user': current_user.to_dict()})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 400

# ========== API Endpoints ==========

@app.route('/api/health', methods=['POST'])
@login_required
def api_health():
    """Predict disease and get fitness plan"""
    try:
        data = request.get_json(force=True)
        
        # Parse symptoms
        symptoms = [normalize(s) for s in data.get('symptoms', [])]
        other = data.get('other', '')
        if other:
            for s in other.split(','):
                s_norm = normalize(s.strip())
                if s_norm in known_symptoms:
                    symptoms.append(s_norm)
                else:
                    m = match_symptom(s_norm)
                    if m:
                        symptoms.append(m)
        
        if not symptoms:
            return jsonify({'error': 'no symptoms provided'}), 400
        
        # Parse user data
        age = int(data.get('age', 25))
        gender = data.get('gender', 'Male')
        height_cm = float(data.get('height_cm', 170))
        weight_kg = float(data.get('weight_kg', 70))
        goal = data.get('goal', 'stay fit')
        hypertension = data.get('hypertension', 'no')
        name = data.get('name', 'Anonymous')
        
        # Predict
        pred, score, desc, prec = predict_disease(symptoms)
        diet, exercises, rec = get_fitness_plan(age, gender, height_cm, weight_kg, goal, hypertension)
        bmi = round(weight_kg / ((height_cm / 100) ** 2), 2)
        
        return jsonify({
            'name': name,
            'age': age,
            'gender': gender,
            'height_cm': height_cm,
            'weight_kg': weight_kg,
            'bmi': bmi,
            'goal': goal,
            'hypertension': hypertension,
            'symptoms': symptoms,
            'disease': pred,
            'risk_score': score,
            'description': desc,
            'precautions': prec,
            'diet': diet,
            'exercises': exercises,
            'recommendation': rec
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/api/admin/status', methods=['GET'])
def admin_status():
    """Get model and data status"""
    return jsonify({
        'model_accuracy': f'{model_accuracy * 100:.2f}%',
        'known_symptoms': len(known_symptoms),
        'diseases': len(dataset['Disease'].unique()),
        'total_records': len(dataset),
        'fitness_records': len(workout)
    })

@app.route('/api/admin/symptoms', methods=['GET'])
def admin_symptoms():
    """List all known symptoms"""
    return jsonify({'symptoms': sorted(list(known_symptoms))})

@app.route('/api/admin/diseases', methods=['GET'])
def admin_diseases():
    """List all diseases"""
    diseases = dataset['Disease'].unique().tolist()
    return jsonify({'diseases': diseases})

@app.route('/api/admin/add-symptom', methods=['POST'])
def admin_add_symptom():
    """Add new symptom to severity dict"""
    data = request.get_json(force=True)
    sym = normalize(data.get('symptom', ''))
    weight = float(data.get('weight', 1))
    
    if not sym:
        return jsonify({'error': 'symptom required'}), 400
    
    severity_dict[sym] = weight
    return jsonify({'message': f'Symptom {sym} added', 'weight': weight})

@app.route('/api/admin/update-symptom/<sym>', methods=['PUT'])
def admin_update_symptom(sym):
    """Update symptom weight"""
    sym = normalize(sym)
    data = request.get_json(force=True)
    weight = float(data.get('weight', 1))
    
    if sym not in severity_dict:
        return jsonify({'error': 'symptom not found'}), 404
    
    severity_dict[sym] = weight
    return jsonify({'message': f'Symptom {sym} updated', 'weight': weight})

@app.route('/api/admin/predictions/<disease>', methods=['GET'])
def admin_disease_info(disease):
    """Get disease prediction info"""
    try:
        desc = description.loc[description['Disease'] == disease, 'Description'].values[0]
        prec = precaution.loc[
            precaution['Disease'] == disease,
            ['Precaution_1', 'Precaution_2', 'Precaution_3']
        ].values[0]
        prec_list = [p for p in prec if pd.notnull(p)]
        return jsonify({
            'disease': disease,
            'description': desc,
            'precautions': prec_list
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 404

@app.route('/', methods=['GET'])
def serve_dashboard():
    """Serve dashboard HTML"""
    return send_from_directory('.', 'dashboard.html')

@app.route('/<path:path>', methods=['GET'])
def serve_static(path):
    """Serve static files"""
    if os.path.exists(path):
        return send_from_directory('.', path)
    return jsonify({'error': 'not found'}), 404

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
        print("✅ Database initialized")

    print("\n🚀 HealthifyAI API Server Starting...")
    print(f"📍 Visit http://localhost:5000 for the dashboard")
    app.run(host='0.0.0.0', port=5000, debug=False)
